import openpyxl
from openpyxl.worksheet.filters import (
    FilterColumn,
    # Filters,
    CustomFilter,
    CustomFilters,
)

wb_name = "first_wb.xlsx"
ws_name = "Conditional"
wb = openpyxl.load_workbook(wb_name)
ws = wb[ws_name]

# filter column A
# irange = f"A1:A{ws.max_row}"
# ws.auto_filter.ref = irange

# filter all column
ws.auto_filter.ref = ws.dimensions

# programmatic filter
# col = FilterColumn(colId=0)  # for column A
# col.filters = Filters(filter=[263])
# ws.auto_filter.filterColumn.append(col)
# ws.auto_filter.add_sort_condition(f"A1:A{ws.max_row}")

# custom filter
flt1 = CustomFilter(operator="lessThan", val=30)
flt2 = CustomFilter(operator="greaterThan", val=90)

cfs = CustomFilters(customFilter=[flt1])
# apply to **third** column in the range
col = FilterColumn(colId=2, customFilters=cfs)
ws.auto_filter.filterColumn.append(col)

# save file
wb.save(wb_name)
