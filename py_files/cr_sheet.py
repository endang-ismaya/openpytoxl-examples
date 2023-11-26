import openpyxl

# initiate a workbook variable
wb = openpyxl.Workbook()

# create sheets
sheet1 = wb.create_sheet("First")
sheet1 = wb.create_sheet("Second")
sheet1 = wb.create_sheet("Third")

# save the workbook
wb.save("first_wb.xlsx")
