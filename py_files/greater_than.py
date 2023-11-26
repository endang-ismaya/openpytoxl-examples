import openpyxl
from openpyxl.styles import PatternFill
from random import randint

wb_name = "first_wb.xlsx"
ws_name = "Conditional"
wb = openpyxl.load_workbook(wb_name)

if ws_name not in wb.sheetnames:
    wb.create_sheet(ws_name)

ws = wb[ws_name]

rows = [("rand1", "rand2", "rand3", "rand4", "rand5")]
for i in range(50):
    r1 = randint(1, 1001)
    r2 = randint(1, 1001)
    r3 = randint(1, 1001)
    r4 = randint(1, 1001)
    r5 = randint(1, 1001)

    rows.append((r1, r2, r3, r4, r5))


for row in rows:
    ws.append(row)


for row in ws.iter_rows(
    min_row=2, max_row=len(rows), min_col=1, max_col=len(rows[0])
):
    for cell in row:
        if cell.value is not None and cell.value < 100:
            cell.fill = PatternFill(
                start_color="96EFFF", end_color="C5FFF8", fill_type="solid"
            )


# save file
wb.save(wb_name)
