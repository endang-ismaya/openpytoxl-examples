import openpyxl
from openpyxl.styles import PatternFill, Font

"""
Fill Type:
none
solid
gray125
gray0625
darkDown
darkGray
darkGrid
darkHorizontal
darkTrellis
darkUp
darkVertical
lightDown
ligthUp
"""


# defining workbook
wb_name = "first_wb.xlsx"
wb = openpyxl.load_workbook(wb_name)
sh = wb["Text_Alignment"]

# cr obj
bg1 = PatternFill(
    start_color="B31312",  # foreground
    end_color="6B240C",  # background
    fill_type="solid",
)
font1 = Font(name="Berlin Sans FB Demi", color="FDF7E4")
sh["J6"] = "Background Style 2"
sh["J6"].fill = bg1
sh["J6"].font = font1

# save file
wb.save(wb_name)
