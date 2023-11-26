import openpyxl

wb_name = "first_wb.xlsx"
ws_name = "Conditional"
wb = openpyxl.load_workbook(wb_name)

if ws_name not in wb.sheetnames:
    wb.create_sheet(ws_name)

ws = wb[ws_name]

ws["F2"] = "=SUM(A2:E2)"
ws["G2"] = "=AVERAGE(A2:E2)"

# save file
wb.save(wb_name)
