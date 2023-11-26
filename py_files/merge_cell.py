import openpyxl

wb_name = "first_wb.xlsx"
wb = openpyxl.load_workbook(wb_name)
sh = wb["Second"]

# merge
sh.merge_cells("E1:J1")
sh["E1"] = "Merge Cells"

# un-merge
# sh.unmerge_cells("E1:J1")

# save file
wb.save(wb_name)
