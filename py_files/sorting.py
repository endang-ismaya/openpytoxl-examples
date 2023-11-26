import openpyxl

wb_name = "first_wb.xlsx"
ws_name = "Conditional"
wb = openpyxl.load_workbook(wb_name)
ws = wb[ws_name]

# data rows
data_rows = list(ws.iter_rows(min_row=2, values_only=True))
# print(data_rows)

# sort data_rows berdasarkan column ke-3 atau array ke-2
sorted_data = sorted(data_rows, key=lambda item: item[0])
# print(sorted_data)

# clear existing data in the sheet
ws.delete_rows(2, ws.max_row)

# append the row
for row in sorted_data:
    ws.append(row)

# save file
wb.save(wb_name)
