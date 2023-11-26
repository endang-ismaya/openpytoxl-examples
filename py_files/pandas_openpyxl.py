import openpyxl
import pandas as pd

wb_name = "first_wb.xlsx"
ws_name = "family_data"
wb = openpyxl.load_workbook(wb_name)
ws = wb[ws_name]

df = pd.read_excel(wb_name, sheet_name=ws_name)
age = df["age"]

print(f"sum: {age.sum()}")  # 113
print(f"mean: {age.mean():.2f}")  # 16.142857142857142

age_sum = age.sum()
age_avg = age.mean()

ws["E2"] = "Sum of Age:"
ws["F2"] = age_sum

ws["E3"] = "Avg of Age:"
ws["F3"] = age_avg

wb.save(wb_name)
