import openpyxl
from openpyxl.chart import (
    BarChart,
    AreaChart,
    LineChart,
    BubbleChart,
    Reference,
)
from copy import deepcopy

wb_name = "first_wb.xlsx"
ws_name = "family_data"
wb = openpyxl.load_workbook(wb_name)
ws = wb[ws_name]

# ages = [cell.value for cell in ws["B"] if str(cell.value).isdigit()]
# names = [
#     cell.value
#     for cell in ws["A"]
#     if str(cell.value).casefold() != "name".casefold()
# ]

# print(ages)
# print(names)
# data = list(zip(names, ages))
names = [cell.value for cell in ws["A"]]
ages = [cell.value for cell in ws["B"]]
data = list(zip(names, ages))

# cr chart obj
colchart = BarChart()
colchart.type = "col"
colchart.title = "Family Age"
colchart.y_axis.title = "Age"
colchart.x_axis.title = "Name"
# colchart.height = 20
# colchart.width = 30

data = Reference(ws, min_col=2, max_col=2, min_row=1, max_row=len(data))
cats = Reference(ws, min_col=1, max_col=1, min_row=2, max_row=len(data))

# True, jadi series diambil dari row pertama data,
# makanya data reference min_row=1
colchart.add_data(data, titles_from_data=True)
colchart.set_categories(cats)


# Horizontal Chart
chart2 = deepcopy(colchart)
chart2.style = 11
chart2.type = "bar"
chart2.title = "Family Age"

# Line Chart
linechart = LineChart()
linechart.add_data(data=data, titles_from_data=True)
linechart.set_categories(cats)
linechart.title = "Family Age"

# Area Chart
areachart = AreaChart()
areachart.add_data(data=data, titles_from_data=True)
areachart.set_categories(cats)
areachart.title = "Family Age"

# Bubble Chart
bbchart = BubbleChart()
bbchart.add_data(data=data, titles_from_data=True)
bbchart.set_categories(cats)
bbchart.title = "Family Age"

# add chart to cell
ws.add_chart(colchart, "I2")
ws.add_chart(chart2, "R2")
ws.add_chart(linechart, "I18")
ws.add_chart(areachart, "R18")
ws.add_chart(bbchart, "I34")


# save file
wb.save(wb_name)
